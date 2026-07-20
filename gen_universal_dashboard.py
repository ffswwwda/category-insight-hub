#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用维度看板生成器
=================
复用 de-dashboard.html 的 CSS 外壳（视觉一致），内核换成通用维度引擎：
- MODULES 配置驱动分布面板（pill 云 / 横条图）
- 点击药丸/横条 = 钻取筛选（toggle）
- 同维度多选 = 并集(OR)；跨维度 = 交集(AND)
- 保存视图自动命名（去 prompt）+ 购物车指纹跟踪 + 面包屑定位
用法：
  python3 gen_universal_dashboard.py
读取两个 xlsx，输出：
  search-volume-dashboard.html
  feedback-intel-dashboard.html
"""
import openpyxl, json, math, re

SRC_SEARCH = "/Users/fsw/Downloads/4、倒模智能体关键词搜索量数据.xlsx"
SRC_FEEDBACK = "/Users/fsw/Downloads/8、情报系统-用户反馈类情报汇总.xlsx"
OUT_DIR = "/Users/fsw/WorkBuddy/2026-07-15-08-51-27"

# ----------------------------------------------------------------------------
# 工具
# ----------------------------------------------------------------------------
def num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    s = re.sub(r'[^\d.\-]','',str(v))
    if s in ('','-','.'): return None
    try: return float(s)
    except: return None

def split_tags(s):
    if not s: return []
    return [t.strip() for t in re.split(r'[,，/、]', str(s)) if t.strip()]

# ----------------------------------------------------------------------------
# 读取：搜索量数据
# ----------------------------------------------------------------------------
def load_search():
    wb = openpyxl.load_workbook(SRC_SEARCH, data_only=True)
    ws = wb['数据表']
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h else '' for h in rows[0]]
    idx = {h:i for i,h in enumerate(hdr)}
    def g(row,k): return row[idx.get(k,'')] if k in idx else None
    out = []
    for r in rows[1:]:
        if not r or all(c is None for c in r): continue
        kw = g(r,'关键词'); 
        if kw is None: continue
        vol = num(g(r,'月搜索量')) or 0
        buy = num(g(r,'购买率'))
        supply = num(g(r,'需供比'))
        comp = num(g(r,'商品数'))
        # 月搜索量档位
        if vol < 1000: vol_b='<1k'
        elif vol < 10000: vol_b='1k–1万'
        elif vol < 100000: vol_b='1万–10万'
        else: vol_b='>10万'
        # 需供比档位（机会度，越大越蓝海）
        if supply is None: sup_b='未知'
        elif supply < 10: sup_b='<10 红海'
        elif supply < 50: sup_b='10–50'
        elif supply < 200: sup_b='50–200'
        else: sup_b='>200 蓝海'
        # 购买率档位
        if buy is None: br_b='未知'
        elif buy < 0.005: br_b='<0.5%'
        elif buy < 0.02: br_b='0.5%–2%'
        else: br_b='>2%'
        # 商品数档位（竞争度）
        if comp is None: comp_b='未知'
        elif comp < 1000: comp_b='<1k'
        elif comp < 10000: comp_b='1k–1万'
        else: comp_b='>1万'
        out.append({
            'kw': str(kw),
            'kw_cn': str(g(r,'关键词翻译') or ''),
            'type': str(g(r,'流量词类型') or '未标注'),
            'vol': vol,
            'buy': buy,
            'supply': supply,
            'comp': comp,
            'ppc': g(r,'PPC竞价'),
            'vol_b': vol_b,
            'sup_b': sup_b,
            'br_b': br_b,
            'comp_b': comp_b,
        })
    wb.close()
    return out

# ----------------------------------------------------------------------------
# 读取：反馈情报（主表 join 明细）
# ----------------------------------------------------------------------------
def load_feedback():
    wb = openpyxl.load_workbook(SRC_FEEDBACK, data_only=True)
    # 主表
    ws1 = wb['用户反馈情报主表1']
    r1 = list(ws1.iter_rows(values_only=True))
    h1 = [str(h).strip() if h else '' for h in r1[0]]
    i1 = {h:i for i,h in enumerate(h1)}
    def g1(row,k): return row[i1.get(k,'')] if k in i1 else None
    main = {}
    for r in r1[1:]:
        if not r or all(c is None for c in r): continue
        iid = g1(r,'情报ID')
        if iid is None: continue
        main[str(iid)] = {
            'intelType': str(g1(r,'情报类型') or '未标注'),
            'dept': str(g1(r,'部门') or '未标注'),
            'market': str(g1(r,'市场标签') or ''),
            'platform': str(g1(r,'平台标签') or ''),
            'title': str(g1(r,'情报标题') or ''),
        }
    # 明细
    ws2 = wb['反馈事实明细表1']
    r2 = list(ws2.iter_rows(values_only=True))
    h2 = [str(h).strip() if h else '' for h in r2[0]]
    i2 = {h:i for i,h in enumerate(h2)}
    def g2(row,k): return row[i2.get(k,'')] if k in i2 else None
    out = []
    for r in r2[1:]:
        if not r or all(c is None for c in r): continue
        fid = g2(r,'反馈事实ID')
        if fid is None: continue
        iid = str(g2(r,'关联情报ID') or '')
        p = main.get(iid, {})
        out.append({
            'fid': str(fid),
            'intelId': iid,
            'intelTitle': p.get('title',''),
            'intelType': p.get('intelType','未标注'),
            'dept': p.get('dept','未标注'),
            'market': p.get('market',''),
            'platform': p.get('platform',''),
            'product': str(g2(r,'产品|SKU') or ''),
            'theme': str(g2(r,'反馈内容主题') or ''),
            'nature': str(g2(r,'内容性质') or ''),
            'tendency': str(g2(r,'反馈倾向') or '未判断'),
            'need': str(g2(r,'涉及问题|需求') or ''),
            'suggest': str(g2(r,'产品改进建议') or ''),
            'excerpt': str(g2(r,'原始反馈摘录') or ''),
            'std': str(g2(r,'标准化反馈事实') or ''),
        })
    wb.close()
    return out

# ----------------------------------------------------------------------------
# HTML 模板
# ----------------------------------------------------------------------------
TPL = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
  :root{
    --bg:#f6f7f9; --panel:rgba(255,255,255,.72); --panel-2:rgba(255,255,255,.5); --border:rgba(0,0,0,.08);
    --ink:#2b2f33; --ink-2:#5b6470; --muted:#8b94a3; --line:rgba(0,0,0,.05);
    --accent:#00C3F3; --accent-weak:#E6F7FC; --accent-2:#a855f7;
    --rad-grad:linear-gradient(135deg,#00d4ff 0%,#a855f7 50%,#ff6b9d 100%);
    --warn:#e8833a; --ok:#1f9d6b; --danger:#e05656; --star:#f5a623;
    --shadow:0 1px 3px rgba(0,0,0,.04),0 6px 20px rgba(0,0,0,.05);
    --glass-blur:blur(14px); --radius:14px; --radius-sm:10px;
    --font:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Hiragino Sans GB","Microsoft YaHei",sans-serif;
    --mono:"SF Mono",ui-monospace,"JetBrains Mono",Menlo,Consolas,monospace;
  }
  body.dark{
    --bg:#0a0e1a; --panel:rgba(20,28,48,.6); --panel-2:rgba(30,40,66,.45); --border:rgba(255,255,255,.1);
    --ink:#e6ebf5; --ink-2:#aab4c8; --muted:#8893a8; --line:rgba(255,255,255,.08);
    --accent-weak:rgba(0,195,243,.14);
    --shadow:0 1px 3px rgba(0,0,0,.3),0 8px 28px rgba(0,0,0,.4);
    background-image:
      radial-gradient(circle at 18% 82%, rgba(0,195,243,.14) 0%, transparent 45%),
      radial-gradient(circle at 82% 18%, rgba(168,85,247,.12) 0%, transparent 45%),
      linear-gradient(rgba(0,195,243,.06) 1px,transparent 1px),
      linear-gradient(90deg,rgba(0,195,243,.06) 1px,transparent 1px);
  }
  *{box-sizing:border-box}
  html{scroll-behavior:smooth}
  body{margin:0;background-color:var(--bg);color:var(--ink);font-family:var(--font);font-size:14px;line-height:1.55;-webkit-font-smoothing:antialiased;
    background-image:
      radial-gradient(circle at 18% 82%, rgba(0,195,243,.08) 0%, transparent 45%),
      radial-gradient(circle at 82% 18%, rgba(168,85,247,.07) 0%, transparent 45%),
      linear-gradient(rgba(0,195,243,.04) 1px,transparent 1px),
      linear-gradient(90deg,rgba(0,195,243,.04) 1px,transparent 1px);
    background-size:auto,auto,60px 60px,60px 60px;background-attachment:fixed}
  ::-webkit-scrollbar{width:10px;height:10px}
  ::-webkit-scrollbar-thumb{background:#d3d8e0;border-radius:8px;border:3px solid var(--bg)}
  ::-webkit-scrollbar-thumb:hover{background:#bcc3cf}
  .app{display:flex;min-height:100vh}
  .sidebar{width:264px;flex:0 0 264px;background:rgba(255,255,255,.74);backdrop-filter:var(--glass-blur);-webkit-backdrop-filter:var(--glass-blur);border-right:1px solid var(--border);display:flex;flex-direction:column;position:sticky;top:0;height:100vh}
  body.dark .sidebar{background:rgba(20,28,48,.6)}
  .brand{display:flex;gap:11px;align-items:center;padding:18px 18px 14px}
  .brand-mark{width:36px;height:36px;border-radius:10px;background:var(--rad-grad);color:#fff;display:grid;place-items:center;font-size:15px;font-weight:800;box-shadow:0 6px 18px rgba(0,195,243,.32)}
  .brand-name{font-weight:700;font-size:15px;letter-spacing:.2px}
  .brand-sub{font-size:11px;color:var(--muted);letter-spacing:.6px;margin-top:1px}
  .src-card{margin:10px 14px;background:var(--panel-2);border:1px solid var(--border);border-radius:12px;padding:13px 14px}
  .src-card .tag{display:inline-block;background:#fff4e0;color:var(--warn);font-weight:600;border-radius:20px;padding:2px 9px;font-size:10.5px;margin-bottom:8px}
  .src-card h4{margin:0 0 4px;font-size:13.5px;font-weight:700;line-height:1.4}
  .src-card .meta{font-size:11.5px;color:var(--muted);line-height:1.6}
  .src-card .meta b{color:var(--ink-2);font-weight:600}
  .view-layers{margin:10px 14px;border:1px solid var(--border);border-radius:12px;background:var(--panel-2);overflow:hidden}
  .vl-head{display:flex;align-items:center;gap:7px;padding:10px 13px;font-size:11px;font-weight:800;color:var(--muted);letter-spacing:.4px;text-transform:uppercase;border-bottom:1px solid var(--line)}
  .vl-head .vl-count{margin-left:auto;background:var(--rad-grad);color:#fff;font-size:9.5px;padding:1px 7px;border-radius:8px;font-weight:700}
  .vl-list{padding:6px;max-height:300px;overflow-y:auto;display:flex;flex-direction:column;gap:4px}
  .vl-item{display:flex;align-items:center;gap:8px;padding:8px 10px;border-radius:9px;border:1px solid transparent;cursor:pointer;transition:all .15s;background:var(--panel)}
  .vl-item:hover{border-color:rgba(0,195,243,.3);background:rgba(0,195,243,.05)}
  .vl-item.active{border-color:var(--accent);background:var(--accent-weak);box-shadow:0 0 0 1px rgba(0,195,243,.15)}
  .vl-item .vl-dot{width:8px;height:8px;border-radius:50%;flex-shrink:0;background:var(--rad-grad)}
  .vl-item .nm{flex:1;font-size:11.5px;font-weight:600;color:var(--ink);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  .vl-item .vl-meta{font-size:9.5px;color:var(--muted);flex-shrink:0}
  .vl-item .vl-del{opacity:0;font-size:13px;color:var(--muted);cursor:pointer;flex-shrink:0;padding:2px 4px;border-radius:4px}
  .vl-item:hover .vl-del{opacity:.6}
  .vl-item .vl-del:hover{opacity:1;color:var(--danger)}
  .vl-empty{color:var(--muted);font-size:11.5px;text-align:center;padding:18px 8px;line-height:1.6}
  .save-view-btn{display:inline-flex;align-items:center;gap:5px;font-size:11.5px;font-weight:700;padding:5px 13px;border:1px solid var(--accent);border-radius:9px;background:var(--accent-weak);color:var(--accent);cursor:pointer;transition:all .15s;white-space:nowrap}
  .save-view-btn:hover{background:var(--accent);color:#fff}
  #cartBtn.cart-added{border-color:#10b981;background:#ecfdf5;color:#059669;cursor:default;pointer-events:none;opacity:.8}
  body.dark #cartBtn.cart-added{background:rgba(16,185,129,.18);color:#5fd6a8}
  .main{flex:1;min-width:0;display:flex;flex-direction:column}
  .topbar{position:sticky;top:0;z-index:20;background:rgba(255,255,255,.72);backdrop-filter:var(--glass-blur);-webkit-backdrop-filter:var(--glass-blur);border-bottom:1px solid var(--border);padding:16px 26px;display:flex;align-items:center;gap:18px}
  body.dark .topbar{background:rgba(20,28,48,.6)}
  .topbar-left{display:flex;flex-direction:column;min-width:0}
  .topbar-titlerow{display:flex;align-items:center;gap:13px}
  .back-btn{appearance:none;border:1px solid var(--border);background:var(--panel);color:var(--ink-2);font-family:var(--font);font-size:12.5px;font-weight:600;padding:7px 12px;border-radius:9px;cursor:pointer;transition:.15s;white-space:nowrap}
  .back-btn:hover{background:var(--accent-weak);border-color:var(--accent);color:var(--accent)}
  .board-title{margin:0;font-size:18px;font-weight:700;letter-spacing:.2px}
  .board-sub{font-size:12px;color:var(--muted);margin-top:2px}
  .topbar-right{margin-left:auto;display:flex;align-items:center;gap:10px}
  .search-box{display:flex;align-items:center;gap:7px;background:var(--panel);border:1px solid var(--border);border-radius:10px;padding:8px 12px;width:320px}
  .search-box input{border:0;outline:0;background:transparent;font-size:13px;width:100%;color:var(--ink);font-family:var(--font)}
  .search-box .ico{color:var(--accent)}
  .fs-btn{appearance:none;display:inline-flex;align-items:center;gap:5px;border:1px solid var(--border);background:var(--panel);color:var(--ink-2);font-family:var(--font);font-size:12px;font-weight:600;padding:7px 11px;border-radius:9px;cursor:pointer;transition:.15s;white-space:nowrap}
  .fs-btn:hover{background:var(--accent-weak);border-color:var(--accent);color:var(--accent)}
  .board{flex:1;padding:22px 26px 60px;overflow:auto}
  .filterbar{background:var(--panel);backdrop-filter:var(--glass-blur);-webkit-backdrop-filter:var(--glass-blur);border:1px solid var(--border);border-radius:var(--radius);padding:14px 16px;margin-bottom:18px;box-shadow:var(--shadow);position:sticky;top:60px;z-index:19}
  .fb-hint{font-size:12px;color:var(--muted);margin-bottom:10px}
  .fb-hint code{background:var(--line);padding:1px 6px;border-radius:5px;font-family:var(--mono);font-size:11.5px}
  .chips{display:flex;flex-wrap:wrap;gap:7px}
  .chip{display:inline-flex;align-items:center;gap:5px;border:1px solid var(--border);background:var(--panel-2);border-radius:20px;padding:5px 12px;font-size:12.5px;cursor:pointer;color:var(--ink-2);transition:.15s;user-select:none}
  .chip:hover{border-color:#c4cad3}
  .chip.on{background:var(--rad-grad);border-color:transparent;color:#fff;box-shadow:0 4px 14px rgba(0,195,243,.3)}
  .fb-foot{display:flex;align-items:center;gap:12px;margin-top:12px;flex-wrap:wrap}
  .clearbtn{margin-left:auto;font-size:12.5px;color:var(--muted);cursor:pointer;text-decoration:underline}
  .understood{font-size:12px;color:var(--accent);font-weight:600;margin-top:8px;min-height:16px}
  .countline{font-size:13px;color:var(--ink-2);margin:0 2px 14px}
  .countline b{color:var(--accent);font-size:15px;font-weight:700}
  .grid{display:grid;gap:16px}
  .g-3{grid-template-columns:repeat(3,1fr)}
  .g-4{grid-template-columns:repeat(4,1fr)}
  .g-6{grid-template-columns:repeat(6,1fr)}
  @media(max-width:1280px){.g-6{grid-template-columns:repeat(3,1fr)}.g-4{grid-template-columns:repeat(2,1fr)}}
  @media(max-width:760px){.g-3,.g-4,.g-6{grid-template-columns:repeat(2,1fr)}.search-box{width:200px}}
  .stat{background:var(--panel);border:1px solid var(--border);border-radius:var(--radius);padding:15px 17px;box-shadow:var(--shadow);transition:.18s;position:relative;overflow:hidden}
  .stat::before{content:'';position:absolute;left:0;top:14px;bottom:14px;width:3px;border-radius:3px;background:var(--rad-grad)}
  .stat:hover{border-color:rgba(0,195,243,.4);box-shadow:0 4px 10px rgba(20,30,50,.05),0 14px 34px rgba(0,195,243,.14)}
  .stat .n{font-size:25px;font-weight:700;letter-spacing:.3px;font-variant-numeric:tabular-nums}
  .stat .l{font-size:12px;color:var(--muted);margin-top:3px}
  .stat .s{font-size:11px;color:var(--accent);margin-top:6px;font-weight:600}
  .panel{background:var(--panel);border:1px solid var(--border);border-radius:var(--radius);padding:18px 20px;box-shadow:var(--shadow);margin-bottom:16px;position:relative;overflow:hidden}
  .panel::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:var(--rad-grad);opacity:.85}
  .panel h3{margin:0 0 14px;font-size:15px;font-weight:700;display:flex;align-items:center;gap:8px}
  .panel h3 .sub{font-size:11.5px;color:var(--muted);font-weight:500;margin-left:auto}
  .dist-cols{display:grid;grid-template-columns:1fr 1fr;gap:24px}
  @media(max-width:900px){.dist-cols{grid-template-columns:1fr}}
  .dist-row{display:flex;align-items:center;gap:12px;padding:6px 0;font-size:13px;cursor:pointer;border-radius:8px;transition:.12s}
  .dist-row:hover{background:var(--panel-2)}
  .dist-row.on{background:var(--accent-weak);box-shadow:inset 0 0 0 1px var(--accent)}
  .dist-row .lab{width:170px;flex:0 0 170px;color:var(--ink-2);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  .dist-row .track{flex:1;height:9px;background:var(--line);border-radius:6px;overflow:hidden}
  .dist-row .track .fill{display:block;height:100%;background:var(--rad-grad);border-radius:6px;transition:width .3s}
  .dist-row .num{width:64px;text-align:right;color:var(--muted);font-variant-numeric:tabular-nums}
  .pill-cloud{display:flex;flex-wrap:wrap;gap:8px;padding:4px 0}
  .pill-tag{display:inline-flex;align-items:center;gap:5px;background:var(--accent-weak);color:var(--accent);border:1px solid rgba(0,195,243,.2);padding:5px 11px;border-radius:16px;font-size:12.5px;transition:.12s;cursor:pointer;user-select:none}
  .pill-tag b{font-weight:700;font-variant-numeric:tabular-nums}
  .pill-tag:hover{background:rgba(0,195,243,.18);transform:scale(1.03)}
  .pill-tag.on{background:var(--rad-grad);border-color:transparent;color:#fff;box-shadow:0 4px 14px rgba(0,195,243,.3)}
  body.dark .pill-tag{background:rgba(0,195,243,.14)}
  .table-wrap{border:1px solid var(--border);border-radius:12px;overflow:hidden}
  table{width:100%;border-collapse:collapse;font-size:12.5px}
  thead th{background:var(--panel-2);text-align:left;padding:10px 12px;font-weight:600;color:var(--ink-2);border-bottom:1px solid var(--border);position:sticky;top:0;white-space:nowrap}
  tbody td{padding:9px 12px;border-bottom:1px solid var(--line);color:var(--ink-2);vertical-align:top}
  tbody tr{cursor:pointer;transition:background .12s}
  tbody tr:hover{background:var(--panel-2)}
  tbody tr.open{background:var(--accent-weak)}
  .ttitle{color:var(--ink);font-weight:500;line-height:1.4;max-width:340px}
  .empty{padding:50px 20px;text-align:center;color:var(--muted)}
</style>
</head>
<body>
<script>
(function(){
  function applyTheme(t){ t=(t==='dark')?'dark':'light'; document.body.classList.toggle('dark',t==='dark'); try{localStorage.setItem('cih-theme',t);}catch(e){} }
  var q=new URLSearchParams(location.search).get('theme'); var s=null; try{s=localStorage.getItem('cih-theme');}catch(e){}
  applyTheme(q||s||'light');
  window.addEventListener('message',function(e){ if(e.data&&e.data.type==='cih-theme') applyTheme(e.data.value); });
})();
</script>
<div class="app">
  <aside class="sidebar">
    <div class="brand">
      <div class="brand-mark">__MARK__</div>
      <div><div class="brand-name">类目数据看板</div><div class="brand-sub">CATEGORY INSIGHT</div></div>
    </div>
    <div class="src-card">
      <span class="tag">真实数据</span>
      <h4>__SRC_TITLE__</h4>
      <div class="meta">__SRC_META__</div>
    </div>
    <div class="view-layers" id="viewLayers">
      <div class="vl-head"><span>视图图层</span><span class="vl-count" id="vlCount">0</span></div>
      <div class="vl-list" id="vlList"><div class="vl-empty">筛选后点「保存为新视图」<br>即可在此保存多组筛选条件</div></div>
    </div>
    <div class="sidebar-foot" style="margin-top:auto;padding:13px 16px 16px;border-top:1px solid var(--border);font-size:11.5px;color:var(--muted);line-height:1.5">点击分布图中的药丸/横条可钻取筛选；<br>同维度多选为并集，跨维度为交集。</div>
  </aside>
  <main class="main">
    <header class="topbar">
      <div class="topbar-left">
        <div class="topbar-titlerow">
          <button class="back-btn" onclick="goBack()" title="返回类目洞察中枢">← 返回</button>
          <div><h1 class="board-title">__BOARD_TITLE__</h1><div class="board-sub">__BOARD_SUB__</div></div>
        </div>
      </div>
      <div class="topbar-right">
        <button class="fs-btn" id="fsBtn" title="全屏（ESC 退出）"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M8 3H5a2 2 0 0 0-2 2v3"/><path d="M21 8V5a2 2 0 0 0-2-2h-3"/><path d="M3 16v3a2 2 0 0 0 2 2h3"/><path d="M16 21h3a2 2 0 0 0 2-2v-3"/></svg>全屏</button>
        <div class="search-box"><span class="ico">⌕</span><input id="nl-input" type="text" placeholder="搜索关键词 / 内容…" autocomplete="off"></div>
      </div>
    </header>
    <div class="board">
      <div class="filterbar">
        <div class="fb-hint">当前定位：<code id="fbTitle">全部数据</code> · 点击下方任意分布图中的元素即可钻取筛选（可叠加）。</div>
        <div class="chips" id="chips"></div>
        <div class="fb-foot">
          <span class="clearbtn" id="clear-btn">清除全部筛选</span>
          <button class="save-view-btn" id="saveViewBtn"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width:15px;height:15px;vertical-align:-2px;margin-right:4px"><path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2z"/><path d="M17 21v-8H7v8M7 3v5h8"/></svg>保存为新视图</button>
          <button class="save-view-btn" id="cartBtn" title="把当前筛选结果加入购物车"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width:15px;height:15px;vertical-align:-2px;margin-right:4px"><circle cx="9" cy="20" r="1.3"/><circle cx="17" cy="20" r="1.3"/><path d="M2 3h2.4l2.1 12.3a1.6 1.6 0 0 0 1.6 1.3h9.1a1.6 1.6 0 0 0 1.6-1.3L21.5 7H5.4"/></svg>加入购物车</button>
        </div>
        <div class="understood" id="understood"></div>
      </div>
      <div class="countline">当前命中 <b id="hit-count">0</b> 条（共 <span id="total-count">0</span> 条）</div>
      <div class="grid g-6 section-gap" id="kpis"></div>
      <div class="panel">
        <h3>多维分布钻取 <span class="sub">点击元素筛选 · 同维度并集 / 跨维度交集</span></h3>
        <div class="dist-cols" id="mod-cols"></div>
      </div>
      <div class="panel">
        <h3>原始数据明细 <span class="sub">点击任意行展开详情</span></h3>
        <div class="table-wrap"><table><thead><tr id="raw-head"></tr></thead><tbody id="raw-body"></tbody></table></div>
      </div>
      <div class="empty" id="empty-state" style="display:none">没有符合当前筛选条件的数据，试试放宽条件。</div>
    </div>
  </main>
</div>
<script>
var ROWS = __ROWS__;
var MODULES = __MODULES__;
var KPIS = __KPIS__;
var COLS = __COLS__;
var SAVED_VIEWS = [];
var ACTIVE_VIEW_ID = null;
var filters = {};
var _cartAdded = false, _cartFp = '';
var curKeyword = '';

function esc(s){ if(s==null) return ''; return String(s).replace(/[&<>"]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];}); }
function fmtNum(n){ if(n==null||isNaN(n)) return '0'; var a=Math.abs(n); if(a>=1e8) return (n/1e8).toFixed(2)+'亿'; if(a>=1e4) return (n/1e4).toFixed(2)+'万'; if(a>=1000) return n.toLocaleString('en-US'); return (''+n); }
function fmtPct(n){ if(n==null||isNaN(n)) return '0%'; return (n*100).toFixed(2)+'%'; }

function matchRow(r){
  for(var i=0;i<MODULES.length;i++){
    var m=MODULES[i]; var set=filters[m.key];
    if(set && set.size){
      var raw=r[m.field]; if(raw==null) return false;
      var vals = m.split ? (''+raw).split(/[,，/、]/).map(function(t){return t.trim();}).filter(Boolean) : [''+raw];
      if(!vals.some(function(v){return set.has(v);})) return false;
    }
  }
  if(curKeyword){
    var hay = (r.__search||'').toLowerCase();
    if(hay.indexOf(curKeyword.toLowerCase())<0) return false;
  }
  return true;
}
function getFiltered(){ return ROWS.filter(matchRow); }

function generateViewFingerprint(){
  var parts=[];
  for(var i=0;i<MODULES.length;i++){ var m=MODULES[i]; var set=filters[m.key]; if(set&&set.size){ set.forEach(function(v){ parts.push(m.label+'='+v); }); } }
  if(curKeyword) parts.push('搜索:'+curKeyword);
  return parts.length?parts.join(' + '):'全部数据（无筛选）';
}
function toast(msg){
  var t=document.getElementById('rdToast');
  if(!t){ t=document.createElement('div'); t.id='rdToast'; t.style.cssText='position:fixed;top:18px;left:50%;transform:translateX(-50%);z-index:9999;padding:8px 18px;border-radius:10px;font-size:12.5px;font-weight:600;color:#fff;background:rgba(0,0,0,.76);backdrop-filter:blur(8px);pointer-events:none;opacity:0;transition:opacity .25s'; document.body.appendChild(t); }
  t.textContent=msg; t.style.opacity='1'; clearTimeout(toast._tm); toast._tm=setTimeout(function(){ t.style.opacity='0'; },2000);
}

function drillToggle(modKey, val){
  filters[modKey]=filters[modKey]||new Set();
  if(filters[modKey].has(val)){ filters[modKey].delete(val); if(filters[modKey].size===0) delete filters[modKey]; }
  else filters[modKey].add(val);
  ACTIVE_VIEW_ID=null; _cartAdded=false; _cartFp=''; updateCartBtn(); render();
}
function clearAll(){
  filters={}; curKeyword=''; ACTIVE_VIEW_ID=null; _cartFp=''; _cartAdded=false;
  document.getElementById('nl-input').value=''; updateCartBtn(); render();
}
function setKeyword(k){ curKeyword=k; ACTIVE_VIEW_ID=null; _cartFp=''; _cartAdded=false; render(); }

function renderKPIs(arr){
  var box=document.getElementById('kpis');
  box.innerHTML=KPIS.map(function(k){
    var v;
    if(k.calc==='count') v=arr.length;
    else if(k.calc==='sum') v=arr.reduce(function(a,r){return a+(+r[k.field]||0);},0);
    else if(k.calc==='avg') v=arr.length?arr.reduce(function(a,r){return a+(+r[k.field]||0);},0)/arr.length:0;
    else if(k.calc==='max') v=arr.reduce(function(a,r){return Math.max(a,+(r[k.field]||0));},0);
    else if(k.calc==='distinct'){ var s=new Set(); arr.forEach(function(r){ (r[k.field]||'').split(/[,，/、]/).forEach(function(t){t=t.trim(); if(t) s.add(t);}); }); v=s.size; }
    else if(k.calc==='where'){ v=arr.filter(function(r){ var x=r[k.field]; if(k.op==='>') return (+x||0)>k.val; if(k.op==='<') return (+x||0)<k.val; if(k.op==='nonempty') return x!=null&&(''+x).trim()!==''; if(k.op==='contains') return (''+x).indexOf(k.val)>=0; if(k.op==='eq') return (''+x)===('')+k.val; return false; }).length; }
    else v=0;
    if(k.fmt==='pct') v=fmtPct(v); else if(k.fmt==='num') v=fmtNum(v); else v=fmtNum(v);
    return '<div class="stat"><div class="n">'+v+'</div><div class="l">'+esc(k.label)+'</div></div>';
  }).join('');
}

function renderModules(arr){
  MODULES.forEach(function(m){
    var el=document.getElementById('mod-'+m.key); if(!el) return;
    var counts={};
    arr.forEach(function(r){ var raw=r[m.field]; if(raw==null||raw==='') return; var vals = m.split ? (''+raw).split(/[,，/、]/).map(function(t){return t.trim();}).filter(Boolean) : [''+raw]; vals.forEach(function(v){ counts[v]=(counts[v]||0)+1; }); });
    var entries=Object.keys(counts).map(function(k){return [k,counts[k]];});
    if(m.order) entries.sort(function(a,b){ return m.order.indexOf(a[0])-m.order.indexOf(b[0]); });
    else entries.sort(function(a,b){ return b[1]-a[1]; });
    var set=filters[m.key];
    function tag(v,c){ var on=set&&set.has(v); return '<span class="pill-tag'+(on?' on':'')+'" onclick="drillToggle(\''+m.key+'\',\''+esc(v).replace(/'/g,"\\'")+'\')">'+esc(v)+' <b>'+c.toLocaleString()+'</b></span>'; }
    function bar(v,c,max){ var on=set&&set.has(v); var pct=max?c/max*100:0; return '<div class="dist-row'+(on?' on':'')+'" onclick="drillToggle(\''+m.key+'\',\''+esc(v).replace(/'/g,"\\'")+'\')"><span class="lab" title="'+esc(v)+'">'+esc(v)+'</span><span class="track"><span class="fill" style="width:'+pct+'%"></span></span><span class="num">'+c.toLocaleString()+'</span></div>'; }
    var inner;
    if(m.kind==='pill' || (entries.length<=10 && entries.every(function(e){return (''+e[0]).length<=16;}))){
      inner='<div class="pill-cloud">'+entries.map(function(e){return tag(e[0],e[1]);}).join('')+'</div>';
    } else {
      var max=entries.length?entries[0][1]:1;
      inner=entries.slice(0,m.top||14).map(function(e){return bar(e[0],e[1],max);}).join('');
    }
    el.innerHTML=inner;
  });
}

function renderChips(){
  var box=document.getElementById('chips');
  var html='';
  MODULES.forEach(function(m){
    var set=filters[m.key]; if(!set) return;
    set.forEach(function(v){ html+='<span class="chip on" onclick="drillToggle(\''+m.key+'\',\''+esc(v).replace(/'/g,"\\'")+'\')">'+esc(m.label)+': '+esc(v)+' ✕</span>'; });
  });
  if(curKeyword) html+='<span class="chip on" onclick="setKeyword(\'\')">搜索: '+esc(curKeyword)+' ✕</span>';
  box.innerHTML=html;
}

function applyFmt(c,r){ var v=r[c.k]; if(c.fmt==='num') return fmtNum(v); if(c.fmt==='pct') return fmtPct(v); return v; }
function renderRaw(arr){
  var head=document.getElementById('raw-head');
  head.innerHTML=COLS.map(function(c){return '<th>'+esc(c.label)+'</th>';}).join('')+'<th></th>';
  var body=document.getElementById('raw-body');
  var slice=arr.slice(0,300);
  body.innerHTML=slice.map(function(r){
    var tds=COLS.map(function(c){ if(c.k==='__search') return ''; return '<td'+(c.cls?' class="'+c.cls+'"':'')+'>'+esc(applyFmt(c,r))+'</td>'; }).join('');
    var det=COLS.map(function(c){ if(c.k==='__search') return ''; return '<div class="kv"><span class="k">'+esc(c.label)+'</span><span class="v">'+esc(applyFmt(c,r))+'</span></div>'; }).join('');
    return '<tr class="mainrow" onclick="var d=this.nextElementSibling;d.style.display=d.style.display===\'none\'?\'table-row\':\'none\';this.classList.toggle(\'open\')">'+tds+'<td style="width:28px;color:var(--muted);text-align:center">▸</td></tr>'+
           '<tr style="display:none"><td colspan="'+(COLS.length+1)+'" style="padding:0"><div class="detail-inner" style="padding:14px 18px;display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:8px 24px">'+det+'</div></td></tr>';
  }).join('');
}

function render(){
  var arr=getFiltered();
  document.getElementById('hit-count').textContent=arr.length.toLocaleString();
  document.getElementById('total-count').textContent=ROWS.length.toLocaleString();
  document.getElementById('fbTitle').textContent=(curKeyword?'搜索: '+curKeyword:(Object.keys(filters).length?'已筛选':'全部数据'));
  renderKPIs(arr);
  renderModules(arr);
  renderChips(arr);
  renderRaw(arr);
  document.getElementById('empty-state').style.display=arr.length?'none':'block';
}

function saveCurrentView(){
  var fp=generateViewFingerprint();
  var dup=SAVED_VIEWS.find(function(v){return v.fp===fp;});
  if(dup){ toast('该筛选组合已存在：'+dup.name); loadView(dup.id); return; }
  var fcopy={}; for(var k in filters){ fcopy[k]=[...filters[k]]; }
  var view={id:Date.now(),name:fp,filters:fcopy,keyword:curKeyword,hitCount:getFiltered().length,fp:fp,createdAt:new Date().toLocaleTimeString('zh-CN',{hour:'2-digit',minute:'2-digit'})};
  SAVED_VIEWS.push(view); try{localStorage.setItem('__LSKEY__',JSON.stringify(SAVED_VIEWS));}catch(e){}
  ACTIVE_VIEW_ID=view.id; renderSavedViews(); toast('已保存视图：'+view.name);
}
function loadView(id){
  var v=SAVED_VIEWS.find(function(x){return x.id===id;}); if(!v) return;
  filters={}; for(var k in (v.filters||{})) filters[k]=new Set(v.filters[k]);
  curKeyword=v.keyword||''; document.getElementById('nl-input').value=curKeyword;
  ACTIVE_VIEW_ID=id; _cartAdded=false; _cartFp=''; updateCartBtn(); render();
  document.querySelectorAll('#vlList .vl-item').forEach(function(el){ el.classList.toggle('active',+el.dataset.vid===id); });
}
function renderSavedViews(){
  var box=document.getElementById('vlList');
  document.getElementById('vlCount').textContent=SAVED_VIEWS.length;
  if(!SAVED_VIEWS.length){ box.innerHTML='<div class="vl-empty">筛选后点「保存为新视图」<br>即可在此保存多组筛选条件</div>'; return; }
  box.innerHTML=SAVED_VIEWS.map(function(v){
    var act=(v.id===ACTIVE_VIEW_ID)?' active':'';
    var nm=v.name.length>30?v.name.slice(0,29)+'…':v.name;
    return '<div class="vl-item'+act+'" data-vid="'+v.id+'" role="button" tabindex="0"><span class="vl-dot"></span><span class="nm" title="'+esc(v.name)+'">'+esc(nm)+'</span><span class="vl-meta">'+v.hitCount.toLocaleString()+' 条</span><span class="vl-del" onclick="delView(event,'+v.id+')">✕</span></div>';
  }).join('');
  box.onclick=function(e){ if(e.target.closest('.vl-del')) return; var it=e.target.closest('.vl-item'); if(it) loadView(+it.dataset.vid); };
  box.onkeydown=function(e){ if(e.key!=='Enter'&&e.key!==' ') return; var it=e.target.closest('.vl-item'); if(it){ e.preventDefault(); loadView(+it.dataset.vid); } };
}
function delView(e,id){ e.stopPropagation(); SAVED_VIEWS=SAVED_VIEWS.filter(function(v){return v.id!==id;}); try{localStorage.setItem('__LSKEY__',JSON.stringify(SAVED_VIEWS));}catch(x){} renderSavedViews(); }
function addToCart(){
  if(_cartAdded) return;
  var arr=getFiltered(); if(!arr.length){ toast('当前筛选结果为空，无法加入购物车'); return; }
  var item={ id:'cart_'+Date.now(), source:'__CART_SRC__', sourceKey:'__CART_KEY__', filter:generateViewFingerprint(), count:arr.length, addedAt:new Date().toISOString(), columns:COLS.map(function(c){return c.k;}), rows:arr.map(function(r){ var o={}; COLS.forEach(function(c){ if(c.k!=='__search') o[c.k]=r[c.k]; }); return o; }) };
  try{ parent.postMessage({type:'cih-cart-add',item:item},'*'); }catch(e){}
  _cartAdded=true; _cartFp=generateViewFingerprint(); updateCartBtn();
  if(window.parent===window){
    var key='cih_cart_v1'; var a=[]; try{a=JSON.parse(localStorage.getItem(key))||[];}catch(e){}
    var d=a.find(function(x){return x.source===item.source&&x.filter===item.filter;}); if(d){d.rows=item.rows;d.count=item.count;d.addedAt=item.addedAt;} else a.push(item);
    try{localStorage.setItem(key,JSON.stringify(a));}catch(e){}
    toast('已加入购物车（'+arr.length+' 条）');
  } else toast('已发送到购物车');
}
function updateCartBtn(){
  var btn=document.getElementById('cartBtn'); var cur=generateViewFingerprint();
  if(_cartAdded && _cartFp===cur){ btn.className='save-view-btn cart-added'; btn.innerHTML='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width:15px;height:15px;vertical-align:-2px;margin-right:4px"><path d="M20 6L9 17l-5-5"/></svg>已加入'; }
  else { if(_cartFp!==cur){_cartFp=cur;_cartAdded=false;} btn.className='save-view-btn'; btn.innerHTML='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width:15px;height:15px;vertical-align:-2px;margin-right:4px"><circle cx="9" cy="20" r="1.3"/><circle cx="17" cy="20" r="1.3"/><path d="M2 3h2.4l2.1 12.3a1.6 1.6 0 0 0 1.6 1.3h9.1a1.6 1.6 0 0 0 1.6-1.3L21.5 7H5.4"/></svg>加入购物车'; }
}
function goBack(){ try{ if(window.parent!==window) window.parent.postMessage({type:'cih-go-back'},'*'); }catch(e){} history.back(); }
function toggleFs(){ var el=document.documentElement; if(!document.fullscreenElement){ el.requestFullscreen&&el.requestFullscreen(); } else { document.exitFullscreen&&document.exitFullscreen(); } }

// 构建模块面板
(function(){
  var cols=document.getElementById('mod-cols');
  var half=Math.ceil(MODULES.length/2);
  var left='', right='';
  MODULES.forEach(function(m,idx){
    var panel='<div style="margin-bottom:18px"><div class="dist-row" style="color:var(--muted);font-size:12px;font-weight:700;cursor:default"><span class="lab">'+esc(m.label)+'</span><span class="track"></span><span class="num">占比</span></div><div id="mod-'+m.key+'"></div></div>';
    if(idx<half) left+=panel; else right+=panel;
  });
  cols.innerHTML='<div>'+left+'</div><div>'+right+'</div>';
})();

// 预计算搜索串
ROWS.forEach(function(r){ r.__search=COLS.map(function(c){return r[c.k];}).filter(Boolean).join(' '); });

// 事件绑定
document.getElementById('saveViewBtn').onclick=saveCurrentView;
document.getElementById('cartBtn').onclick=addToCart;
document.getElementById('clear-btn').onclick=clearAll;
document.getElementById('fsBtn').onclick=toggleFs;
document.getElementById('nl-input').addEventListener('input',function(e){ setKeyword(e.target.value.trim()); });

// 还原视图
try{ var sv=JSON.parse(localStorage.getItem('__LSKEY__')||'[]'); if(Array.isArray(sv)) SAVED_VIEWS.push.apply(SAVED_VIEWS,sv); }catch(e){}
renderSavedViews();
updateCartBtn();
render();
</script>
</body>
</html>
"""

def build(rows, modules, kpis, cols, meta, title, board_title, board_sub, mark, lskey, cart_src, cart_key, outfile):
    # build module columns container already in TPL
    html = TPL
    html = html.replace('__TITLE__', title)
    html = html.replace('__MARK__', mark)
    html = html.replace('__SRC_TITLE__', meta['title'])
    html = html.replace('__SRC_META__', meta['meta'])
    html = html.replace('__BOARD_TITLE__', board_title)
    html = html.replace('__BOARD_SUB__', board_sub)
    html = html.replace('__ROWS__', json.dumps(rows, ensure_ascii=False))
    html = html.replace('__MODULES__', json.dumps(modules, ensure_ascii=False))
    html = html.replace('__KPIS__', json.dumps(kpis, ensure_ascii=False))
    html = html.replace('__COLS__', json.dumps(cols, ensure_ascii=False))
    html = html.replace('__LSKEY__', lskey)
    html = html.replace('__CART_SRC__', cart_src)
    html = html.replace('__CART_KEY__', cart_key)
    with open(outfile,'w') as f:
        f.write(html)
    print('wrote', outfile, len(html), 'bytes', '| rows', len(rows), '| modules', len(modules))

# ----------------------------------------------------------------------------
# 搜索量看板
# ----------------------------------------------------------------------------
def main():
    search = load_search()
    search_cols = [
        {'k':'kw','label':'关键词','cls':'ttitle'},
        {'k':'kw_cn','label':'翻译'},
        {'k':'type','label':'词类型'},
        {'k':'vol','label':'月搜索量','fmt':'num'},
        {'k':'buy','label':'购买率','fmt':'pct'},
        {'k':'supply','label':'需供比','fmt':'num'},
        {'k':'comp','label':'商品数','fmt':'num'},
        {'k':'ppc','label':'PPC竞价'},
    ]
    search_modules = [
        {'key':'type','label':'流量词类型','field':'type','kind':'pill'},
        {'key':'kw_cn','label':'关键词翻译（Top 15）','field':'kw_cn','kind':'bar','top':15},
        {'key':'vol_b','label':'月搜索量档位','field':'vol_b','kind':'bar','order':['<1k','1k–1万','1万–10万','>10万']},
        {'key':'sup_b','label':'需供比档位（蓝海机会）','field':'sup_b','kind':'bar','order':['<10 红海','10–50','50–200','>200 蓝海','未知']},
        {'key':'br_b','label':'购买率档位','field':'br_b','kind':'bar','order':['<0.5%','0.5%–2%','>2%','未知']},
        {'key':'comp_b','label':'商品数档位（竞争度）','field':'comp_b','kind':'bar','order':['<1k','1k–1万','>1万','未知']},
    ]
    search_kpis = [
        {'label':'关键词总数','calc':'count'},
        {'label':'总月搜索量','calc':'sum','field':'vol','fmt':'num'},
        {'label':'平均购买率','calc':'avg','field':'buy','fmt':'pct'},
        {'label':'总商品数（竞争池）','calc':'sum','field':'comp','fmt':'num'},
        {'label':'蓝海词（需供比>200）','calc':'where','field':'supply','op':'>','val':200},
        {'label':'高购买率词（>2%）','calc':'where','field':'buy','op':'>','val':0.02},
    ]
    # 注意：后两个 KPI 是 count 全部行，需改为条件计数——简化为展示总量，下面用 JS 不支持条件；改为 distinct/数值近似
    build(
        rows=search, modules=search_modules, kpis=search_kpis, cols=search_cols,
        meta={'title':'倒模智能体 · 亚马逊关键词搜索量','meta':'数据规模：<b>'+str(len(search))+'</b> 个关键词<br>覆盖：<b>亚马逊全站点</b><br>维度：搜索量/购买率/需供比/竞争<br>来源：卖家精灵关键词库'},
        title='倒模智能体 · 亚马逊关键词搜索量看板',
        board_title='倒模智能体 · 亚马逊关键词搜索量看板',
        board_sub=str(len(search))+' 个真实关键词 · 流量词类型 / 搜索量 / 需供比 / 竞争度多维钻取',
        mark='KW', lskey='kw_views', cart_src='亚马逊关键词搜索量', cart_key='search',
        outfile=OUT_DIR+'/vendor/search-volume-dashboard.html'
    )

    # ----------------------------------------------------------------------------
    # 反馈情报看板
    # ----------------------------------------------------------------------------
    fb = load_feedback()
    fb_cols = [
        {'k':'fid','label':'事实ID'},
        {'k':'intelTitle','label':'关联情报','cls':'ttitle'},
        {'k':'product','label':'产品|SKU'},
        {'k':'theme','label':'反馈主题'},
        {'k':'tendency','label':'反馈倾向'},
        {'k':'std','label':'标准化事实'},
        {'k':'need','label':'涉及问题/需求'},
        {'k':'suggest','label':'改进建议'},
    ]
    fb_modules = [
        {'key':'intelType','label':'情报类型','field':'intelType','kind':'pill'},
        {'key':'dept','label':'提交部门','field':'dept','kind':'pill'},
        {'key':'market','label':'市场标签（Top 15）','field':'market','kind':'bar','top':15,'split':True},
        {'key':'platform','label':'平台标签','field':'platform','kind':'pill'},
        {'key':'product','label':'产品|SKU（Top 15）','field':'product','kind':'bar','top':15,'split':True},
        {'key':'theme','label':'反馈内容主题（Top 15）','field':'theme','kind':'bar','top':15,'split':True},
        {'key':'tendency','label':'反馈倾向','field':'tendency','kind':'pill'},
        {'key':'nature','label':'内容性质','field':'nature','kind':'pill'},
    ]
    fb_kpis = [
        {'label':'反馈事实总数','calc':'count'},
        {'label':'关联情报数','calc':'distinct','field':'intelId'},
        {'label':'涉及产品数','calc':'distinct','field':'product'},
        {'label':'覆盖市场数','calc':'distinct','field':'market'},
        {'label':'含改进建议','calc':'where','field':'suggest','op':'nonempty'},
        {'label':'含问题/需求','calc':'where','field':'need','op':'nonempty'},
    ]
    build(
        rows=fb, modules=fb_modules, kpis=fb_kpis, cols=fb_cols,
        meta={'title':'情报系统 · 用户反馈类情报汇总','meta':'数据规模：<b>'+str(len(fb))+'</b> 条反馈事实<br>关联情报：<b>87</b> 份<br>涵盖：访谈/问卷/社媒风评<br>来源：情报系统导出'},
        title='情报系统 · 用户反馈类情报汇总看板',
        board_title='情报系统 · 用户反馈类情报汇总看板',
        board_sub=str(len(fb))+' 条反馈事实（来自 87 份情报）· 情报类型 / 产品 / 市场 / 主题多维钻取',
        mark='FB', lskey='fb_views', cart_src='用户反馈情报', cart_key='feedback',
        outfile=OUT_DIR+'/vendor/feedback-intel-dashboard.html'
    )

if __name__=='__main__':
    main()
