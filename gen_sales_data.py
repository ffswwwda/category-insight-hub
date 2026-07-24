import openpyxl, json, re

SRC = "/Users/fsw/Downloads/2、商品销售数据【美亚+日亚+欧亚】.xlsx"
OUT = "/Users/fsw/WorkBuddy/2026-07-15-08-51-27/sales_data_new.json"

wb = openpyxl.load_workbook(SRC, read_only=False, data_only=True)

def colmap(name):
    ws = wb[name]
    m = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is not None:
            m[h] = c
    return ws, m

def g(ws, m, row, *names):
    """return first non-None value among candidate header names"""
    for n in names:
        if n in m:
            v = ws.cell(row=row, column=m[n]).value
            if v is not None:
                return v
    return None

def gpre(ws, m, row, prefix):
    for h, c in m.items():
        if isinstance(h, str) and h.startswith(prefix):
            v = ws.cell(row=row, column=c).value
            if v is not None:
                return v
    return None

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s in ('', 'None', 'N/A', '#N/A', '-', '—'): return None
    s = re.sub(r'[^\d.\-]', '', s)
    if s in ('', '.', '-'): return None
    try: return float(s)
    except: return None

def clean_str(v):
    if v is None: return None
    s = str(v).replace('\n', ' ').replace('\r', ' ').strip()
    if s in ('None', 'N/A', '#N/A', '-', '—', ''): return None
    # collapse whitespace
    s = re.sub(r'\s+', ' ', s)
    return s

def trim(v, n):
    if v is None: return None
    v = str(v)
    if len(v) > n:
        v = v[:n].rstrip() + '…'
    return v

def split_tags(v):
    if v is None: return []
    s = str(v)
    parts = re.split(r'[，,、/|;；\n]+', s)
    out = []
    for p in parts:
        p = p.strip()
        if p and p not in ('None', 'N/A', '#N/A', '-', '未提及', ''):
            out.append(p)
    # dedupe preserve order
    seen = set(); res = []
    for p in out:
        if p not in seen:
            seen.add(p); res.append(p)
    return res

# ---- structured field header names (master) ----
STRUCT = {
    'mat': '材质材料', 'matTouch': '材质触感', 'channel': '通道设计',
    'texture': '内部纹理', 'feature': '功能特点', 'size': '尺寸重量',
    'scene': '使用场景', 'portability': '便携性', 'realism': '逼真仿真度',
    'cleaning': '清洁维护', 'privacy': '隐私包装', 'cert': '安全认证',
    'aud': '适用人群', 'category': '产品品类', 'semantic': '语义理解',
}

def month_label(sheet):
    m = re.search(r'(\d{4})年(\d{1,2})月', sheet)
    if m:
        return "%s-%02d" % (m.group(1), int(m.group(2)))
    return sheet

# ---------- MASTER parse ----------
def parse_master(sheet):
    ws, m = colmap(sheet)
    out = {}
    for r in range(2, ws.max_row + 1):
        asin = clean_str(g(ws, m, r, 'ASIN'))
        if not asin:
            continue
        rec = {
            'asin': asin,
            'brand': clean_str(g(ws, m, r, '品牌')) or '',
            'title': clean_str(g(ws, m, r, '商品标题')) or '',
            'zhTitle': clean_str(g(ws, m, r, '标题翻译')),
            'sp': trim(clean_str(g(ws, m, r, '产品卖点')), 350),
            'zhSp': trim(clean_str(g(ws, m, r, '卖点翻译')), 350),
            'cat': clean_str(g(ws, m, r, '大类目')),
            'sub': clean_str(g(ws, m, r, '小类目')),
            'catPath': clean_str(g(ws, m, r, '类目路径')),
            'rating': num(g(ws, m, r, '评分')),
            'reviews': num(g(ws, m, r, '评分数')),
            'margin': num(g(ws, m, r, '毛利率')),
            'shipping': clean_str(g(ws, m, r, '配送方式')),
            'sellerLoc': clean_str(g(ws, m, r, '卖家所属地')),
            'aplus': clean_str(g(ws, m, r, 'A+页面')),
            'video': clean_str(g(ws, m, r, '视频介绍')),
            'origin': clean_str(g(ws, m, r, '产地')) or clean_str(g(ws, m, r, '卖家所属地')),
        }
        for k, h in STRUCT.items():
            rec[k] = trim(clean_str(g(ws, m, r, h)), 150)
        # tags from master if present
        mt = g(ws, m, r, '标签')
        rec['_master_tags'] = split_tags(mt)
        # numeric metrics also possibly in master (DE/UK/JP)
        rec['_m_price'] = num(gpre(ws, m, r, '价格'))
        rec['_m_units'] = num(g(ws, m, r, '月销量'))
        rec['_m_sales'] = num(gpre(ws, m, r, '月销售额'))
        rec['_m_fba'] = num(gpre(ws, m, r, 'FBA'))
        out[asin] = rec
    return out

# ---------- MONTHLY parse ----------
def parse_monthly(sheet):
    ws, m = colmap(sheet)
    ml = month_label(sheet)
    items = []
    tag_union = {}  # asin -> set
    for r in range(2, ws.max_row + 1):
        asin = clean_str(g(ws, m, r, 'ASIN'))
        if not asin:
            continue
        rec = {
            'asin': asin,
            'brand': clean_str(g(ws, m, r, '品牌')) or '',
            'title': clean_str(g(ws, m, r, '商品标题')) or '',
            'sp': trim(clean_str(g(ws, m, r, '产品卖点')), 350),
            'cat': clean_str(g(ws, m, r, '大类目')),
            'sub': clean_str(g(ws, m, r, '小类目')),
            'catPath': clean_str(g(ws, m, r, '类目路径')),
            'price': num(gpre(ws, m, r, '价格')),
            'units': num(g(ws, m, r, '月销量')),
            'sales': num(gpre(ws, m, r, '月销售额')),
            'rating': num(g(ws, m, r, '评分')),
            'reviews': num(g(ws, m, r, '评分数')),
            'fba': num(gpre(ws, m, r, 'FBA')),
            'margin': num(g(ws, m, r, '毛利率')),
            'shipping': clean_str(g(ws, m, r, '配送方式')),
            'sellerLoc': clean_str(g(ws, m, r, '卖家所属地')),
            'aplus': clean_str(g(ws, m, r, 'A+页面')),
            'video': clean_str(g(ws, m, r, '视频介绍')),
        }
        tg = split_tags(g(ws, m, r, '标签'))
        rec['tags'] = tg
        if asin not in tag_union:
            tag_union[asin] = set()
        tag_union[asin].update(tg)
        items.append(rec)
    return ml, items, tag_union

# ---------- CONFIG ----------
CFG = {
    '美亚': {
        'currency': '$', 'mkts': ['us'],
        'master': {'us': '美亚-总表v2'},
        'months': {'us': ['美亚2025年12月','美亚2026年1月','美亚2026年2月','美亚2026年3月','美亚2026年4月','美亚2026年5月']},
        'mktName': {'us': '美亚'},
    },
    '日亚': {
        'currency': '¥', 'mkts': ['jp'],
        'master': {'jp': '日亚总表'},
        'months': {'jp': ['日亚2025年12月','日亚2026年1月','日亚2026年2月','日亚2026年3月','日亚2026年4月','日亚2026年5月']},
        'mktName': {'jp': '日亚'},
    },
    '欧亚（德国+英国）': {
        'currency': '€', 'mkts': ['de', 'uk'],
        'master': {'de': '欧亚-德国总表', 'uk': '欧亚-英国总表'},
        'months': {
            'de': ['欧亚-德国2025年12月','欧亚-德国2026年1月','欧亚-德国2026年2月','欧亚-德国2026年3月','欧亚-德国2026年4月','欧亚-德国2026年5月'],
            'uk': ['欧亚-英国2025年12月','欧亚-英国2026年1月','欧亚-英国2026年2月','欧亚-英国2026年3月','欧亚-英国2026年4月','欧亚-英国2026年5月'],
        },
        'mktName': {'de': '德国', 'uk': '英国'},
    },
}

# parse all monthly sheets once
monthly_cache = {}
for board, bc in CFG.items():
    for mkt, sheets in bc['months'].items():
        for sh in sheets:
            ml, items, tu = parse_monthly(sh)
            monthly_cache.setdefault(board, {}).setdefault(mkt, []).append((ml, items, tu))

# build SALES_DATA
SALES_DATA = {
    'currency': {b: bc['currency'] for b, bc in CFG.items()},
    'boardMeta': {},
    'boards': {},
}

for board, bc in CFG.items():
    SALES_DATA['boardMeta'][board] = {
        'markets': bc['mkts'],
        'mktName': bc['mktName'],
        'months': sorted(set(ml for mkt in bc['months'] for ml, _, _ in monthly_cache[board][mkt])),
    }
    totals = {}
    months_out = {}
    for mkt in bc['mkts']:
        master = parse_master(bc['master'][mkt])
        mdata = monthly_cache[board][mkt]  # list of (ml, items, tu)
        # union tags across months per asin
        union_tags = {}
        for ml, items, tu in mdata:
            for a, s in tu.items():
                union_tags.setdefault(a, set()).update(s)
        # latest month items by asin (for totals numbers)
        latest_ml, latest_items, _ = mdata[-1]
        latest_by_asin = {it['asin']: it for it in latest_items}
        # build totals: from master, merge latest month numbers, tags
        tlist = []
        for asin, rec in master.items():
            nr = latest_by_asin.get(asin, {})
            out = {
                'asin': asin, 'brand': rec['brand'], 'title': rec['title'], 'zhTitle': rec['zhTitle'],
                'sp': rec['sp'], 'zhSp': rec['zhSp'],
                'cat': rec['cat'], 'sub': rec['sub'], 'catPath': rec['catPath'],
                'price': nr.get('price', rec['_m_price']),
                'units': nr.get('units', rec['_m_units']),
                'sales': nr.get('sales', rec['_m_sales']),
                'rating': rec['rating'] if rec['rating'] is not None else nr.get('rating'),
                'reviews': rec['reviews'] if rec['reviews'] is not None else nr.get('reviews'),
                'fba': nr.get('fba', rec['_m_fba']),
                'margin': rec['margin'] if rec['margin'] is not None else nr.get('margin'),
                'shipping': rec['shipping'] or nr.get('shipping'),
                'sellerLoc': rec['sellerLoc'] or nr.get('sellerLoc'),
                'aplus': rec['aplus'] or nr.get('aplus'),
                'video': rec['video'] or nr.get('video'),
                'origin': rec['origin'],
            }
            for k in STRUCT:
                out[k] = rec[k]
            # tags: master 标签 if present, else monthly union; plus meaningful structured attrs
            tg = list(rec['_master_tags']) if rec['_master_tags'] else sorted(union_tags.get(asin, []))
            for k in STRUCT:
                v = rec[k]
                if v and v != '未提及' and v not in tg:
                    tg.append(v)
            out['tags'] = tg
            tlist.append(out)
        totals[mkt] = tlist
        # months out
        mlist = []
        for ml, items, tu in mdata:
            mitems = []
            for it in items:
                a = it['asin']
                base = master.get(a, {})
                o = {
                    'asin': a, 'brand': it['brand'], 'title': it['title'], 'zhTitle': base.get('zhTitle'),
                    'sp': it['sp'], 'zhSp': base.get('zhSp'),
                    'cat': it['cat'], 'sub': it['sub'], 'catPath': it['catPath'],
                    'price': it['price'], 'units': it['units'], 'sales': it['sales'],
                    'rating': it['rating'], 'reviews': it['reviews'], 'fba': it['fba'], 'margin': it['margin'],
                    'shipping': it['shipping'], 'sellerLoc': it['sellerLoc'],
                    'aplus': it['aplus'], 'video': it['video'], 'origin': base.get('origin'),
                }
                for k in STRUCT:
                    o[k] = base.get(k)
                tg = list(it['tags']) if it['tags'] else sorted(union_tags.get(a, []))
                for k in STRUCT:
                    v = base.get(k)
                    if v and v != '未提及' and v not in tg:
                        tg.append(v)
                o['tags'] = tg
                mitems.append(o)
            mlist.append({'month': ml, 'items': mitems})
        months_out[mkt] = mlist
    SALES_DATA['boards'][board] = {'totals': totals, 'months': months_out}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(SALES_DATA, f, ensure_ascii=False, separators=(',', ':'))

# stats
print("WROTE", OUT)
for b, bc in CFG.items():
    for mkt in bc['mkts']:
        t = SALES_DATA['boards'][b]['totals'][mkt]
        tagged = sum(1 for x in t if x['tags'])
        zh = sum(1 for x in t if x['zhTitle'])
        attr = sum(1 for x in t if any(x.get(k) and x.get(k) != '未提及' for k in STRUCT))
        avg = sum(len(x['tags']) for x in t) / max(1, len(t))
        print("  %s/%s totals=%d tagged=%d zhTitle=%d attributed=%d avgTags=%.1f" % (b, mkt, len(t), tagged, zh, attr, avg))
print("TOTAL bytes:", len(json.dumps(SALES_DATA, ensure_ascii=False)))
